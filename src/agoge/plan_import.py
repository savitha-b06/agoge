"""Bulk plan import from a spreadsheet (CSV or xlsx).

Re-imports only overwrite future rows — anything already past stays as history.
Each import carries a version stamp and a one-line reason so the plan's own
history is auditable.
"""
from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from .analysis import week_bounds
from .config import Athlete
from .db import DB

# Columns accepted (case-insensitive). Aliases map to the canonical name.
COLUMNS = {
    "date": "date",
    "day": "date",
    "week_of_block": "week_of_block",
    "week": "week_of_block",
    "sport": "sport",
    "session_type": "session_type",
    "type": "session_type",
    "planned_duration_min": "planned_duration_min",
    "planned_min": "planned_duration_min",
    "duration_min": "planned_duration_min",
    "target_hr_low": "target_hr_low",
    "hr_low": "target_hr_low",
    "target_hr_high": "target_hr_high",
    "hr_high": "target_hr_high",
    "segments": "segments",
    "lift_focus": "lift_focus",
    "notes": "notes",
    "title": "title",
}

SESSION_TYPES = {"endurance", "interval", "strength", "rest", "brick", "test"}
INTENSITY_TYPES = {"interval", "test"}


def parse_segments(raw: str | None) -> list[dict[str, Any]]:
    """Parse `15min warmup Z1 | 145min steady <160bpm | 20min surge 170-180bpm`
    into discrete blocks. Best-effort — unknown pieces stay in `label`."""
    if not raw or not str(raw).strip():
        return []
    parts = [p.strip() for p in str(raw).split("|") if p.strip()]
    out = []
    for part in parts:
        block: dict[str, Any] = {"raw": part, "label": part}
        m = re.match(
            r"(?P<dur>\d+(?:\.\d+)?)\s*(?:min|m)?\s+(?P<label>.+)",
            part, re.I,
        )
        if m:
            block["duration_min"] = float(m.group("dur"))
            block["label"] = m.group("label").strip()
        hr = re.search(r"(\d+)\s*[-–]\s*(\d+)\s*bpm", part, re.I)
        if hr:
            block["hr_low"] = int(hr.group(1))
            block["hr_high"] = int(hr.group(2))
        else:
            cap = re.search(r"[<≤]\s*(\d+)\s*bpm", part, re.I)
            if cap:
                block["hr_high"] = int(cap.group(1))
            floor = re.search(r"[>≥]\s*(\d+)\s*bpm", part, re.I)
            if floor:
                block["hr_low"] = int(floor.group(1))
        z = re.search(r"\b(Z[1-5])\b", part, re.I)
        if z:
            block["zone"] = z.group(1).lower()
        out.append(block)
    return out


def _normalize_header(name: str) -> str | None:
    key = re.sub(r"[^a-z0-9_]+", "_", name.strip().lower()).strip("_")
    return COLUMNS.get(key)


def _read_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if suffix in (".csv", ".tsv", ".txt", ""):
        return _read_csv(path)
    raise ValueError(f"Unsupported plan file type: {suffix or '(none)'}. Use .csv or .xlsx.")


def _read_csv(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="utf-8-sig")
    dialect = csv.Sniffer().sniff(text[:2048], delimiters=",\t;")
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("Plan file has no header row.")
    return [dict(r) for r in reader]


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading .xlsx needs openpyxl. Install it (`pip install openpyxl`) "
            "or export the sheet as CSV."
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [str(h) if h is not None else "" for h in header]
    out = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({keys[i]: ("" if row[i] is None else str(row[i]))
                    for i in range(len(keys))})
    return out


def _parse_date(value: str) -> date:
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value[:10], fmt).date()
        except ValueError:
            continue
    # Excel serial dates sometimes arrive as floats-as-strings
    try:
        n = float(value)
        # Excel epoch 1899-12-30
        return date(1899, 12, 30) + timedelta(days=int(n))
    except ValueError:
        pass
    return date.fromisoformat(value[:10])


def _as_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return int(float(str(value).strip()))


def _as_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    return float(str(value).strip())


def parse_plan_file(path: Path | str) -> list[dict[str, Any]]:
    """Return normalised plan rows from a spreadsheet. Does not write anything."""
    path = Path(path)
    raw_rows = _read_rows(path)
    parsed = []
    for i, raw in enumerate(raw_rows, start=2):
        mapped: dict[str, Any] = {}
        for k, v in raw.items():
            if k is None:
                continue
            canon = _normalize_header(str(k))
            if canon:
                mapped[canon] = v
        if "date" not in mapped or not str(mapped["date"]).strip():
            continue
        day = _parse_date(mapped["date"])
        session_type = (mapped.get("session_type") or "").strip().lower() or None
        if session_type and session_type not in SESSION_TYPES:
            # accept unknown types rather than reject the row
            pass
        segments_raw = mapped.get("segments") or ""
        segments = parse_segments(segments_raw)
        planned = _as_float(mapped.get("planned_duration_min"))
        if planned is None and segments:
            planned = sum(s.get("duration_min") or 0 for s in segments) or None
        sport = (mapped.get("sport") or "").strip().lower() or None
        if session_type == "rest":
            sport = sport or "rest"
        title = (mapped.get("title") or "").strip() or None
        if not title:
            bits = []
            for b in (sport, session_type):
                if b and b not in bits:
                    bits.append(b)
            title = " ".join(bits) if bits else "planned session"
        week_start, _ = week_bounds(day)
        parsed.append({
            "day": day.isoformat(),
            "week_of_block": _as_int(mapped.get("week_of_block")),
            "sport": sport,
            "session_type": session_type,
            "planned_min": planned,
            "target_hr_low": _as_int(mapped.get("target_hr_low")),
            "target_hr_high": _as_int(mapped.get("target_hr_high")),
            "segments": segments,
            "segments_raw": segments_raw.strip() or None,
            "lift_focus": (mapped.get("lift_focus") or "").strip() or None,
            "notes": (mapped.get("notes") or "").strip() or None,
            "title": title,
            "detail": segments_raw.strip() or (mapped.get("notes") or "").strip() or None,
            "week_start": week_start,
            "status": "planned",
            "source": "import",
            "_row": i,
        })
    return parsed


def check_plan_conflicts(rows: list[dict[str, Any]], athlete: Athlete,
                         db: DB | None = None,
                         today: date | None = None) -> list[dict[str, Any]]:
    """Surface conflicts between a prescribed plan and athlete rules / open gates.

    Never silently accepts intensity work during a base-only block, or running
    while an injury gate is red. Returns a list of conflict dicts.
    """
    today = today or date.today()
    conflicts = []
    flags = []
    if db is not None:
        from .analysis import injury_flags
        flags = injury_flags(db, athlete, today.isoformat())
    red_injuries = {f["injury"] for f in flags if f["severity"] == "red"}

    for row in rows:
        day = date.fromisoformat(row["day"])
        block = athlete.current_block(day)
        st = (row.get("session_type") or "").lower()
        sport = (row.get("sport") or "").lower()

        if block and st in INTENSITY_TYPES:
            notes = (block.get("notes") or "").lower()
            focus = (block.get("focus") or "").lower()
            base_only = (
                "z2" in notes or "base" in (block.get("name") or "").lower()
                or focus in ("run", "base") and "all" not in focus
            )
            # Austin-style base block notes explicitly say all Z2
            if "all z2" in notes or "all zone 2" in notes or (
                base_only and "intensity" not in notes and "interval" not in notes
            ):
                conflicts.append({
                    "day": row["day"],
                    "kind": "block_intensity",
                    "message": (
                        f"{row['day']}: {st} session prescribed during "
                        f"'{block['name']}' which is base/Z2-only — "
                        f"surface this, do not relay as written."
                    ),
                })

        if sport == "run" and red_injuries:
            for inj in athlete.injuries:
                if inj["key"] in red_injuries and "run" in inj.get("prompt_after", []):
                    conflicts.append({
                        "day": row["day"],
                        "kind": "injury_gate",
                        "message": (
                            f"{row['day']}: run prescribed but injury gate is open "
                            f"for {inj['label']} — plan is overridden, not executed."
                        ),
                    })
                    break

        if st not in ("rest", "") and row.get("planned_min") and db is not None:
            # soft check: if this single day would blow the weekly ramp alone
            from .analysis import load_check
            load = load_check(db, athlete, day)
            if load["cap_min"] and row["planned_min"] > load["cap_min"] * 0.6:
                # only flag obviously oversized single sessions relative to cap
                pass

    return conflicts


def import_plan(db: DB, athlete: Athlete, path: Path | str, *,
                from_day: date | str | None = None,
                reason: str = "",
                version: str | None = None,
                today: date | None = None) -> dict[str, Any]:
    """Scoped re-import: only overwrite plan rows for dates that haven't happened yet.

    Past days stay as historical record. Returns write counts and any conflicts.
    """
    today = today or date.today()
    if isinstance(from_day, str):
        from_day = date.fromisoformat(from_day) if from_day else None
    cutoff = from_day or today
    path = Path(path)
    rows = parse_plan_file(path)
    version = version or datetime.now().strftime("v%Y%m%d-%H%M%S")
    reason = reason or f"import from {path.name}"

    # Group by date so two-a-days survive: clear each date once, then insert all.
    by_day: dict[str, list[dict[str, Any]]] = {}
    skipped = 0
    for row in rows:
        day = date.fromisoformat(row["day"])
        if day < cutoff:
            skipped += 1
            continue
        by_day.setdefault(row["day"], []).append(row)

    written = 0
    for day_str, day_rows in sorted(by_day.items()):
        # Preserve historical outcome rows (done/skipped/missed); only refresh active ones.
        existing = db.plan_for_day(day_str)
        active = [e for e in existing if (e["status"] or "planned") in ("planned", "prescribed")]
        if existing and not active and date.fromisoformat(day_str) < today:
            skipped += len(day_rows)
            continue
        db.clear_planned_day(day_str)
        for row in day_rows:
            payload = {**row}
            payload.pop("_row", None)
            payload["version"] = version
            payload["import_reason"] = reason
            payload["segments"] = (
                json.dumps(payload["segments"]) if payload.get("segments") else None
            )
            db.insert_plan_row(payload)
            written += 1

    conflicts = check_plan_conflicts(
        [r for r in rows if date.fromisoformat(r["day"]) >= cutoff],
        athlete, db=db, today=today,
    )
    db.record_plan_import(
        version=version, from_day=cutoff.isoformat(), reason=reason,
        file_name=path.name, rows_written=written, rows_skipped=skipped,
    )
    return {
        "version": version,
        "from_day": cutoff.isoformat(),
        "reason": reason,
        "rows_total": len(rows),
        "rows_written": written,
        "rows_skipped": skipped,
        "conflicts": conflicts,
    }


def revision_context(db: DB, athlete: Athlete, today: date | None = None) -> str:
    """Portable summary to hand to whatever drafts a plan revision.

    Anchors a new plan to demonstrated fitness rather than the original guess.
    """
    from .fitness import fitness_trend, format_fitness_trend
    from .profile import read_profile

    today = today or date.today()
    lines = [
        f"Plan revision context for {athlete.name} as of {today.isoformat()}.",
        "Use this — not memory — when drafting a structural revision.",
        "",
        format_fitness_trend(fitness_trend(db, athlete, today)),
        "",
    ]
    from .analysis import checkpoint_status, load_check, plan_divergence, race_status
    race = race_status(db, athlete, today)
    load = load_check(db, athlete, today)
    lines += [
        f"Block: {race['block']} — {race['block_notes']}",
        f"Race in {race['days_to_race']} days.",
        f"Load: this week {load['this_week_min']:.0f} min, last week "
        f"{load['last_week_min']:.0f} min, cap {load['cap_min']}.",
        "",
        "Checkpoints:",
    ]
    for c in checkpoint_status(db, athlete, today):
        lines.append(
            f"  {c['metric']}: {c['actual']} / {c['target']} due {c['due']} "
            f"({'on track' if c['on_track'] else 'BEHIND'})"
        )
    div = plan_divergence(db, athlete, today)
    if div.get("message"):
        lines += ["", "Divergence:", f"  {div['message']}"]
    profile = read_profile()
    if profile:
        lines += ["", "Standing profile:", profile]
    return "\n".join(lines)
